from openpyxl import *

path = r"C:\Users\Emre Can AltınIşık\Desktop\Bionluk\deneme.xlsx"
wb = load_workbook(path)
ws = wb.active
max_column = ws.max_column
max_row = ws.max_row

#Bu "DİĞER" kısmı
for satir in range(2, max_row+1):
    breadcrumb_sutun = 17
    a = str(ws.cell(satir, breadcrumb_sutun, ).value)
    if "Diğer>" in a:
        b = a.replace("Diğer>", "")
        str(ws.cell(satir, breadcrumb_sutun, value=str(b)))

#Yaş Grubu - Varyasyon ilişkisi
for satir in range(2, max_row+1):
    aciklama_sutun = 8
    varyasyon_sutun = 4
    a = str(ws.cell(satir, aciklama_sutun).value)
    if "Yaş Grubu : 3+ Yaş" in a:
        ws.cell(satir, varyasyon_sutun, value="3-5, 6-9, 10-12, 13+")
    elif "Yaş Grubu : 4+ Yaş" in a:
        ws.cell(satir, varyasyon_sutun, value="3-5, 6-9, 10-12, 13+")
    elif "Yaş Grubu : 5+ Yaş" in a:
        ws.cell(satir, varyasyon_sutun, value="3-5, 6-9, 10-12, 13+")
    elif "Yaş Grubu : 6+ Yaş" in a:
        ws.cell(satir, varyasyon_sutun, value="6-9, 10-12, 13+")
    elif "Yaş Grubu : 7+ Yaş" in a:
        ws.cell(satir, varyasyon_sutun, value="6-9, 10-12, 13+")
    elif "Yaş Grubu : 8+ Yaş" in a:
        ws.cell(satir, varyasyon_sutun, value="6-9, 10-12, 13+")
    elif "Yaş Grubu : 9+ Yaş" in a:
        ws.cell(satir, varyasyon_sutun, value="6-9, 10-12, 13+")
    elif "Yaş Grubu : 10+ Yaş" in a:
        ws.cell(satir, varyasyon_sutun, value="10-12, 13+")
    elif "Yaş Grubu : 11+ Yaş" in a:
        ws.cell(satir, varyasyon_sutun, value="10-12, 13+")
    elif "Yaş Grubu : 12+ Yaş" in a:
        ws.cell(satir, varyasyon_sutun, value="10-12, 13+")
    elif "Yaş Grubu : 13+ Yaş" in a:
        ws.cell(satir, varyasyon_sutun, value="13+")
    else:
        ws.cell(satir, varyasyon_sutun, value="")
wb.save(r"C:\Users\Emre Can AltınIşık\Desktop\Bionluk\deneme9.xlsx")
