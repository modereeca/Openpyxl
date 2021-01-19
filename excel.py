from openpyxl import load_workbook
from os import path

excelFilePath = path.join(path.curdir, "original.xlsx")
wb = load_workbook(excelFilePath)
print("Dosya yüklendi")
ws = wb.active
max_row = ws.max_row
